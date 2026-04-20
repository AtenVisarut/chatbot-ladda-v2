"""
Run capability test — asks the REAL bot pipeline 6 questions per product,
scores answers heuristically, exports Excel report.

Usage:
  python scripts/run_capability_test.py --limit 3        # quick 3 products
  python scripts/run_capability_test.py --sample         # representative 10
  python scripts/run_capability_test.py --all            # full 90

Output: reports/capability_test_<timestamp>.xlsx (+ .jsonl cache)
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

# --- set env defaults before importing app modules ---
os.environ.setdefault("ADMIN_PASSWORD", "capability-test-only")
os.environ.setdefault("SECRET_KEY", "capability-test-secret-key-1234567")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
load_dotenv()

# Suppress app noise
import logging
logging.basicConfig(level=logging.WARNING)
for noisy in ("app.services.memory", "app.services.cache", "httpx", "app.services.rag.orchestrator"):
    logging.getLogger(noisy).setLevel(logging.ERROR)


from app.dependencies import supabase_client
from app.services.product.registry import ProductRegistry
from app.services.rag.orchestrator import process_with_agentic_rag

from tests.capability_scorer import (
    build_questions,
    score_product_info, score_pest_crop, score_usage_rate,
    score_moa, score_selling_point, score_comparison,
)


REPORTS_DIR = Path(__file__).resolve().parent.parent / "reports"
REPORTS_DIR.mkdir(exist_ok=True)


# Representative sample (10 products covering all categories)
SAMPLE_PRODUCTS = [
    "ไบเตอร์",         # Insecticide
    "คอนทาฟ",          # Fungicide
    "ออล์สตาร์",        # Herbicide
    "โบว์แลน 285",      # Insecticide (neonicotinoid)
    "ไดยูแมกซ์",        # Herbicide
    "บอมส์ ไวท์",       # Fertilizer/Biostimulant
    "แจ๊ส 50 อีซี",    # Insecticide (organophosphate)
    "พรีดิคท์ 25% เอฟ", # PGR
    "NPK 0-0-60",       # Fertilizer
    "อาร์เทมิส",        # Fungicide
]


# =============================================================================
# Data model
# =============================================================================

@dataclass
class CellResult:
    product: str
    capability: int
    question: str
    answer: str
    score: int
    reason: str
    error: Optional[str] = None
    elapsed_ms: int = 0


@dataclass
class ProductResult:
    product: str
    category: str
    cells: List[CellResult] = field(default_factory=list)

    @property
    def avg_score(self) -> float:
        if not self.cells:
            return 0.0
        return sum(c.score for c in self.cells) / len(self.cells)


# =============================================================================
# Fetch products + choose comparison partner
# =============================================================================

async def fetch_products(names_filter: Optional[List[str]] = None) -> List[dict]:
    res = supabase_client.table("products3").select(
        "product_name, common_name_th, active_ingredient, product_category, "
        "applicable_crops, fungicides, insecticides, herbicides, biostimulant, "
        "pgr_hormones, fertilizer, how_to_use, usage_rate, usage_period, "
        "selling_point, absorption_method, mechanism_of_action, physical_form, "
        "chemical_group_rac, caution_notes, package_size"
    ).execute()
    products = res.data or []
    if names_filter:
        name_set = {n.strip() for n in names_filter}
        products = [p for p in products if p.get("product_name") in name_set]
    return products


def pick_compare_partner(target: dict, all_products: List[dict]) -> Optional[dict]:
    """Pick a product in the same category with a different name for comparison test"""
    target_name = target.get("product_name")
    target_cat = (target.get("product_category") or "").lower()
    # Prefer same category
    same_cat = [p for p in all_products
                if p.get("product_name") != target_name
                and (p.get("product_category") or "").lower() == target_cat]
    if same_cat:
        return same_cat[0]
    # Fall back to any other product
    others = [p for p in all_products if p.get("product_name") != target_name]
    return others[0] if others else None


# =============================================================================
# Query + score
# =============================================================================

async def run_query_with_retry(query: str, max_retries: int = 2) -> tuple[str, Optional[str], int]:
    """Run the RAG pipeline once. Returns (answer, error, elapsed_ms)"""
    start = time.time()
    last_err = None
    for attempt in range(max_retries + 1):
        try:
            r = await process_with_agentic_rag(query, context="", user_id=None)
            elapsed = int((time.time() - start) * 1000)
            return (r.answer or "", None, elapsed)
        except Exception as e:
            last_err = str(e)[:200]
            if attempt < max_retries:
                await asyncio.sleep(2 ** attempt)  # exponential backoff
    elapsed = int((time.time() - start) * 1000)
    return ("", last_err, elapsed)


async def evaluate_product(product: dict, partner: Optional[dict]) -> ProductResult:
    """Run 6 questions for 1 product and score them"""
    questions = build_questions(product, comparison_partner=partner)
    result = ProductResult(
        product=product.get("product_name", "?"),
        category=product.get("product_category", "?"),
    )

    for cap_id, q_text in questions.items():
        ans, err, ms = await run_query_with_retry(q_text)

        if cap_id == 1:
            s = score_product_info(ans, product)
        elif cap_id == 2:
            s = score_pest_crop(ans, product)
        elif cap_id == 3:
            s = score_usage_rate(ans, product)
        elif cap_id == 4:
            s = score_moa(ans, product)
        elif cap_id == 5:
            s = score_selling_point(ans, product)
        elif cap_id == 6 and partner:
            s = score_comparison(ans, product, partner)
        else:
            continue

        result.cells.append(CellResult(
            product=result.product, capability=cap_id,
            question=q_text, answer=ans[:600],
            score=s.score, reason=s.reason,
            error=err, elapsed_ms=ms,
        ))
    return result


# =============================================================================
# Concurrent runner with progress
# =============================================================================

async def run_with_concurrency(
    products: List[dict], all_products: List[dict],
    concurrency: int = 4,
    jsonl_cache: Optional[Path] = None,
) -> List[ProductResult]:
    sem = asyncio.Semaphore(concurrency)
    results: List[ProductResult] = []
    total = len(products)
    done = 0
    start_time = time.time()

    # Load existing cache if any (for resume)
    done_names: set = set()
    if jsonl_cache and jsonl_cache.exists():
        with open(jsonl_cache, encoding="utf-8") as f:
            for line in f:
                try:
                    data = json.loads(line)
                    done_names.add(data["product"])
                    pr = ProductResult(product=data["product"], category=data["category"])
                    pr.cells = [CellResult(**c) for c in data["cells"]]
                    results.append(pr)
                except Exception:
                    continue
        if done_names:
            print(f"🔄 Resume: {len(done_names)} products already cached")

    pending = [p for p in products if p.get("product_name") not in done_names]
    done = len(done_names)

    async def worker(product: dict):
        nonlocal done
        async with sem:
            partner = pick_compare_partner(product, all_products)
            try:
                pr = await evaluate_product(product, partner)
            except Exception as e:
                pr = ProductResult(
                    product=product.get("product_name", "?"),
                    category=product.get("product_category", "?"),
                )
                pr.cells.append(CellResult(
                    product=pr.product, capability=0, question="",
                    answer="", score=0, reason=f"fatal: {e}", error=str(e)[:300],
                ))
            results.append(pr)
            done += 1

            # Append to cache for resume safety
            if jsonl_cache:
                with open(jsonl_cache, "a", encoding="utf-8") as f:
                    f.write(json.dumps({
                        "product": pr.product, "category": pr.category,
                        "cells": [asdict(c) for c in pr.cells],
                    }, ensure_ascii=False) + "\n")

            elapsed = int(time.time() - start_time)
            eta = int(elapsed / max(done - len(done_names), 1) * (total - done))
            pct = done / total * 100
            print(
                f"[{done}/{total}] ({pct:.0f}%) "
                f"{pr.product:<30} avg={pr.avg_score:5.1f}  "
                f"t+{elapsed//60}m{elapsed%60}s  eta~{eta//60}m{eta%60}s"
            )

    await asyncio.gather(*(worker(p) for p in pending))
    return results


# =============================================================================
# Excel export
# =============================================================================

def export_excel(results: List[ProductResult], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    # Shared styles
    header_fill = PatternFill("solid", fgColor="1F6FEB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="top", horizontal="left")
    wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Overall metrics (used on multiple sheets)
    all_scores = [c.score for pr in results for c in pr.cells]
    overall = sum(all_scores) / len(all_scores) if all_scores else 0
    total_products = len(results)
    pass_products = sum(1 for pr in results if pr.avg_score >= 70)

    wb = Workbook()

    # =========================================================================
    # Sheet 1: scores — with overall banner + criteria in headers + feedback col
    # =========================================================================
    ws = wb.active
    ws.title = "scores"

    # --- Row 1-3: Overall banner ---
    ws.merge_cells("A1:J1")
    ws["A1"] = f"📊 Capability Test — Overall Accuracy: {overall:.2f} / 100"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F4C81")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"📦 {total_products} products  ·  "
        f"✅ {pass_products}/{total_products} pass (≥70)  ·  "
        f"❌ {total_products - pass_products} need review"
    )
    ws["A2"].font = Font(size=10, italic=True, color="333333")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = PatternFill("solid", fgColor="E7F1FF")
    ws.row_dimensions[2].height = 20

    # Row 3 empty (spacer)
    ws.row_dimensions[3].height = 8

    # --- Row 4: Column headers with criteria ---
    HEADER_ROW = 4
    headers = [
        ("product", "ชื่อสินค้า", ""),
        ("category", "ประเภท", ""),
        ("cap1", "#1 ข้อมูลสินค้า\n(0-100)",
         "name+thai+%+formulation\nแต่ละอย่าง +25"),
        ("cap2", "#2 โรค/แมลง/พืช\n(0-100)",
         "target ≥2=+50, ≥1=+35\ncrop ≥1=+50"),
        ("cap3", "#3 อัตรา+วิธีใช้\n(0-100)",
         "number match=+50\nverb (ผสม/ฉีด/พ่น)=+50"),
        ("cap4", "#4 MoA / IRAC\n(0-100)",
         "group code match=+100\nnarrative keywords ≥2=+100"),
        ("cap5", "#5 จุดเด่น\n(0-100)",
         "selling_point token\noverlap ≥50%=+100"),
        ("cap6", "#6 เปรียบเทียบ\n(0-100)",
         "both names=+50\n+ differentiator=+50"),
        ("avg", "เฉลี่ย\n(1-100)", "mean(cap1..6)"),
        ("status", "status", "≥70 ✅\n50-69 ⚠️\n<50 ❌"),
        ("feedback", "FEEDBACK\n(ใส่ความเห็นเอง)", "ช่องว่าง — admin กรอกเอง"),
    ]
    # Header (2 lines: label + criteria)
    for col_i, (_, label, criteria) in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=col_i)
        if criteria:
            cell.value = f"{label}\n\n📐 {criteria}"
        else:
            cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_center
        cell.border = Border(bottom=Side(style="medium", color="0F4C81"))
    ws.row_dimensions[HEADER_ROW].height = 75

    # --- Data rows (sorted by avg desc) ---
    results_sorted = sorted(results, key=lambda r: r.avg_score, reverse=True)
    for pr in results_sorted:
        by_cap = {c.capability: c.score for c in pr.cells}
        status_icon = ("✅" if pr.avg_score >= 70
                       else "⚠️" if pr.avg_score >= 50 else "❌")
        ws.append([
            pr.product, pr.category,
            by_cap.get(1, "-"), by_cap.get(2, "-"),
            by_cap.get(3, "-"), by_cap.get(4, "-"),
            by_cap.get(5, "-"), by_cap.get(6, "-"),
            round(pr.avg_score, 1),
            status_icon,
            "",  # feedback column — blank for admin to fill
        ])

    first_data_row = HEADER_ROW + 1
    last_data_row = HEADER_ROW + len(results_sorted)

    # Conditional color on avg (col I) and individual caps (C-H)
    avg_col = "I"
    ws.conditional_formatting.add(
        f"{avg_col}{first_data_row}:{avg_col}{last_data_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["70"],
                   fill=PatternFill("solid", fgColor="C6EFCE")),
    )
    ws.conditional_formatting.add(
        f"{avg_col}{first_data_row}:{avg_col}{last_data_row}",
        CellIsRule(operator="lessThan", formula=["50"],
                   fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    # Light highlight for <70 cap scores
    for col_letter in ("C", "D", "E", "F", "G", "H"):
        ws.conditional_formatting.add(
            f"{col_letter}{first_data_row}:{col_letter}{last_data_row}",
            CellIsRule(operator="lessThan", formula=["70"],
                       fill=PatternFill("solid", fgColor="FFF2E6")),
        )
        ws.conditional_formatting.add(
            f"{col_letter}{first_data_row}:{col_letter}{last_data_row}",
            CellIsRule(operator="lessThan", formula=["50"],
                       fill=PatternFill("solid", fgColor="FFC7CE")),
        )

    # Column widths
    widths = [28, 14, 16, 16, 16, 16, 16, 16, 12, 10, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"C{HEADER_ROW + 1}"  # freeze product + category

    # --------------- Sheet 2: details ---------------
    ws2 = wb.create_sheet("details")
    headers2 = ["product", "capability", "question", "bot_answer",
                "score", "reason", "elapsed_ms", "error"]
    ws2.append(headers2)
    for col_i, _ in enumerate(headers2, 1):
        ws2.cell(row=1, column=col_i).fill = PatternFill("solid", fgColor="1F6FEB")
        ws2.cell(row=1, column=col_i).font = Font(bold=True, color="FFFFFF")

    for pr in results_sorted:
        for c in pr.cells:
            ws2.append([
                c.product, f"#{c.capability}",
                c.question, c.answer,
                c.score, c.reason,
                c.elapsed_ms, c.error or "",
            ])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["C"].width = 45
    ws2.column_dimensions["D"].width = 90
    ws2.column_dimensions["F"].width = 40
    ws2.freeze_panes = "A2"
    # wrap text for long columns
    for row in ws2.iter_rows(min_row=2):
        for col_letter in ("C", "D", "F"):
            cell = ws2[f"{col_letter}{row[0].row}"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[row[0].row].height = 80

    # --------------- Sheet 3: summary ---------------
    ws3 = wb.create_sheet("summary")

    # Overall banner on summary too
    ws3.merge_cells("A1:G1")
    ws3["A1"] = f"🏆 Overall Accuracy: {overall:.2f} / 100  ·  {total_products} products  ·  {len(all_scores)} cells tested"
    ws3["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", fgColor="0F4C81")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    headers3 = ["Capability", "Mean", "Median", "Min", "Max", "Pass (≥70)", "Count"]
    ws3.append(headers3)
    for col_i in range(1, 8):
        cell = ws3.cell(row=2, column=col_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_center

    cap_labels = {
        1: "1: ข้อมูลสินค้า",
        2: "2: โรค/แมลง/พืช",
        3: "3: อัตรา + วิธีใช้",
        4: "4: MoA / IRAC",
        5: "5: จุดเด่น",
        6: "6: เปรียบเทียบ",
    }
    for cap_id, label in cap_labels.items():
        scores = [c.score for pr in results for c in pr.cells if c.capability == cap_id]
        if not scores:
            continue
        scores.sort()
        mean = sum(scores) / len(scores)
        median = scores[len(scores) // 2]
        pass_count = sum(1 for s in scores if s >= 70)
        ws3.append([
            label,
            round(mean, 1),
            median,
            min(scores), max(scores),
            f"{pass_count}/{len(scores)} ({pass_count/len(scores)*100:.0f}%)",
            len(scores),
        ])

    ws3.append([])
    ws3.append(["OVERALL AVERAGE", round(overall, 1), "", "", "", "", len(all_scores)])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=12)
    ws3.cell(row=ws3.max_row, column=2).font = Font(bold=True, size=12, color="0F4C81")
    ws3.cell(row=ws3.max_row, column=1).fill = PatternFill("solid", fgColor="FFF9C4")
    ws3.cell(row=ws3.max_row, column=2).fill = PatternFill("solid", fgColor="FFF9C4")
    for i in range(1, 8):
        ws3.column_dimensions[get_column_letter(i)].width = 20

    # --------------- Sheet 4: failed (score < 70) — add feedback column ---------------
    ws4 = wb.create_sheet("failed")
    ws4.append(["product", "capability", "score", "question", "bot_answer", "reason", "feedback"])
    for col_i in range(1, 8):
        ws4.cell(row=1, column=col_i).fill = PatternFill("solid", fgColor="D93025")
        ws4.cell(row=1, column=col_i).font = Font(bold=True, color="FFFFFF")

    for pr in results_sorted:
        for c in pr.cells:
            if c.score < 70:
                ws4.append([
                    c.product, f"#{c.capability}", c.score,
                    c.question, c.answer[:300], c.reason,
                    "",  # feedback blank
                ])
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["D"].width = 45
    ws4.column_dimensions["E"].width = 80
    ws4.column_dimensions["F"].width = 40
    ws4.column_dimensions["G"].width = 40
    ws4.freeze_panes = "A2"

    # --------------- Sheet 5: criteria — explain scoring rubric ---------------
    ws5 = wb.create_sheet("criteria")
    ws5.merge_cells("A1:E1")
    ws5["A1"] = "📐 เกณฑ์การให้คะแนน (Scoring Rubric)"
    ws5["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws5["A1"].fill = PatternFill("solid", fgColor="0F4C81")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28

    headers5 = ["Cap", "ชื่อ", "คำถามที่ใช้", "DB Fields เทียบ", "เกณฑ์ (คะแนนรวม 100)"]
    ws5.append(headers5)
    for col_i in range(1, 6):
        cell = ws5.cell(row=2, column=col_i)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_center

    criteria_rows = [
        ("#1", "ข้อมูลสินค้า",
         '"{product} สารสำคัญคืออะไร formulation แบบไหน"',
         "product_name\ncommon_name_th\nactive_ingredient",
         "• ชื่อสินค้าปรากฏ → +25\n"
         "• ชื่อสารไทย (common_name_th) ปรากฏ → +25\n"
         "• % จาก active_ingredient ตรง → +25\n"
         "• Formulation (EC/SC/WP/WG/SL/GR/…) → +25"),
        ("#2", "โรค/แมลง/พืช",
         '"{product} ใช้กับพืชอะไรได้บ้าง กำจัดอะไร"',
         "insecticides / fungicides /\nherbicides / biostimulant /\npgr_hormones / fertilizer\n+ applicable_crops",
         "• Target (pest/disease) match:\n"
         "    ≥2 tokens → +50\n"
         "    1 token   → +35\n"
         "    0         → 0\n"
         "• Applicable crop ≥1 → +50\n"
         "หมายเหตุ: แยก 'pestในcrop' format อัตโนมัติ"),
        ("#3", "อัตรา + วิธีใช้",
         '"{product} อัตราใช้เท่าไหร่ ผสมยังไง"',
         "usage_rate\nhow_to_use",
         "• ตัวเลขจาก usage_rate ปรากฏ → +50\n"
         "    (ถ้าไม่ตรงแต่มี number+unit → +25)\n"
         "• Action verb → +50\n"
         "    (ผสม/ฉีด/ฉีดพ่น/พ่น/ราด/หว่าน/\n"
         "     หยด/รด/คราด/คลุก — 'ใช้' ไม่นับ)"),
        ("#4", "MoA / IRAC",
         '"{product} อยู่กลุ่ม IRAC/FRAC/HRAC อะไร"',
         "chemical_group_rac\n(+ mechanism_of_action)",
         "• กรณีมี code (เช่น 3A, C2):\n"
         "    ทุก code match → 100\n"
         "    บางส่วน         → 60\n"
         "    ไม่มีเลย        → 0\n"
         "• กรณี narrative (PGR):\n"
         "    keywords ≥2    → 100\n"
         "    keywords = 1   → 60\n"
         "    keywords = 0   → 0\n"
         "• LLM ตอบ 'ไม่มีข้อมูล' + พูดถึงกลุ่ม → 0"),
        ("#5", "จุดเด่น (Selling Point)",
         '"จุดเด่นของ {product} คืออะไร"',
         "selling_point",
         "• Token overlap จาก selling_point:\n"
         "    ≥50% match → 100\n"
         "    25-49%     → 70\n"
         "    ≥1 token  → 40\n"
         "    0          → 0\n"
         "• LLM บอก 'ไม่มีจุดเด่น' → 0\n"
         "• DB ไม่มี selling_point → 100 (skip)"),
        ("#6", "เปรียบเทียบ",
         '"{a} กับ {b} ต่างกันยังไง"',
         "product_name (x2)\nproduct_category\ncommon_name_th\nmechanism_of_action\nusage_rate",
         "• ทั้ง 2 ชื่อใน answer → +50\n"
         "   (1 ชื่อ → +25)\n"
         "• Differentiator → +50 (check 4 แบบ)\n"
         "    1. Category ต่าง + mention\n"
         "    2. Thai ingredient ของทั้ง 2 ฝั่ง\n"
         "    3. Mechanism keywords ≥2\n"
         "    4. Usage rates ต่างกัน (ตัวเลข)"),
    ]
    for row in criteria_rows:
        ws5.append(row)
        r = ws5.max_row
        for col_i in range(1, 6):
            ws5.cell(row=r, column=col_i).alignment = wrap
        ws5.row_dimensions[r].height = 130

    # Add notes / limitations row
    ws5.append([])
    ws5.append(["⚠️ ข้อจำกัด (Honest Limitations)", "",
                "1. Rule-based scorer: ~85-90% accurate vs human review",
                "",
                "2. ไม่วัด: tone, length, flow, helpfulness, emoji density"])
    ws5.append(["", "", "3. คำถามเป็น template (ไม่ใช่ slang/typo จริงของเกษตรกร)", "",
                "4. Single-turn: ไม่ทดสอบ multi-turn context"])
    ws5.append(["", "", "5. False-negative ~10-15% ถ้า LLM ใช้ synonym", "",
                "6. คะแนน 96+ ≠ bot พร้อม release — ต้อง test เพิ่ม (load, multi-turn, edge)"])

    # Column widths
    ws5.column_dimensions["A"].width = 6
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 42
    ws5.column_dimensions["D"].width = 30
    ws5.column_dimensions["E"].width = 55
    ws5.freeze_panes = "A3"

    wb.save(out_path)


# =============================================================================
# Main
# =============================================================================

async def main():
    parser = argparse.ArgumentParser()
    grp = parser.add_mutually_exclusive_group()
    grp.add_argument("--all", action="store_true", help="run full 90 products")
    grp.add_argument("--sample", action="store_true", help="run 10 representative products")
    grp.add_argument("--limit", type=int, help="run first N products")
    grp.add_argument("--products", type=str, help="comma-separated product names")
    parser.add_argument("--concurrency", type=int, default=4,
                        help="concurrent queries (careful with OpenAI rate limits)")
    parser.add_argument("--resume", type=str,
                        help="resume from existing .jsonl cache file")
    args = parser.parse_args()

    # Load registry
    reg = ProductRegistry.get_instance()
    await reg.load_from_db(supabase_client)

    # Fetch all products
    all_products = await fetch_products()
    print(f"📦 DB has {len(all_products)} products")

    # Select subset
    if args.all:
        targets = all_products
    elif args.sample:
        targets = [p for p in all_products if p.get("product_name") in SAMPLE_PRODUCTS]
    elif args.limit:
        targets = all_products[:args.limit]
    elif args.products:
        wanted = {n.strip() for n in args.products.split(",")}
        targets = [p for p in all_products if p.get("product_name") in wanted]
    else:
        print("No mode selected. Use --all, --sample, --limit N, or --products 'A,B'")
        return

    print(f"🎯 Will test {len(targets)} products × 6 capabilities = "
          f"{len(targets) * 6} queries")
    est_cost = len(targets) * 6 * 0.52  # ~0.52 THB per query
    est_time = len(targets) * 6 * 12 / max(args.concurrency, 1) / 60  # min
    print(f"💰 Estimated cost: ~{est_cost:.0f} THB, time: ~{est_time:.0f} min "
          f"(concurrency={args.concurrency})")
    print()

    # Setup cache + output paths
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    cache_path = (Path(args.resume) if args.resume
                  else REPORTS_DIR / f"capability_test_{ts}.jsonl")
    xlsx_path = REPORTS_DIR / f"capability_test_{ts}.xlsx"

    # Run
    start = time.time()
    results = await run_with_concurrency(
        targets, all_products,
        concurrency=args.concurrency,
        jsonl_cache=cache_path,
    )
    elapsed = int(time.time() - start)
    print(f"\n⏱  Done in {elapsed//60}m{elapsed%60}s")

    # Export
    export_excel(results, xlsx_path)
    print(f"📊 Report saved: {xlsx_path}")
    print(f"💾 Cache (for resume): {cache_path}")

    # Print summary to stdout
    all_scores = [c.score for pr in results for c in pr.cells]
    overall = sum(all_scores) / len(all_scores) if all_scores else 0
    print(f"\n🏆 Overall average: {overall:.1f} / 100")
    for cap_id in (1, 2, 3, 4, 5, 6):
        scores = [c.score for pr in results for c in pr.cells if c.capability == cap_id]
        if scores:
            mean = sum(scores) / len(scores)
            passed = sum(1 for s in scores if s >= 70)
            print(f"   Cap #{cap_id}: {mean:5.1f}  ({passed}/{len(scores)} ≥70)")


if __name__ == "__main__":
    asyncio.run(main())
