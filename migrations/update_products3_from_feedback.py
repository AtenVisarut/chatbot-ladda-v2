# -*- coding: utf-8 -*-
"""
Update products3 from feedback Excel file.
Reads สรุปfeedbackสำหรับทีมPD (1).xlsx and updates Supabase.
"""
import os
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from dotenv import load_dotenv
load_dotenv(override=True)

import openpyxl
from supabase import create_client

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
PRODUCT_TABLE = os.getenv("PRODUCT_TABLE", "products3")
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'สรุปfeedbackสำหรับทีมPD (1).xlsx')


def clean_text(val):
    """Clean cell value: strip, normalize whitespace."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    return s


def read_excel():
    """Read all product updates from Excel, return list of dicts."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    updates = []

    # Sheet config: (sheet_name, pest_column_name, pest_col_index)
    sheet_config = [
        ('Insecticide (ยาแมลง)', 'insecticides', 7),
        ('Fungicide (ยาโรค)', 'fungicides', 7),
        ('Herbicide (ยาหญ้า)', 'herbicides', 7),
        ('PGR (สารควบคุม)', None, None),
    ]

    for sheet_name, pest_col, pest_idx in sheet_config:
        ws = wb[sheet_name]
        for r in range(4, ws.max_row + 1):
            name = clean_text(ws.cell(r, 2).value)
            if not name:
                continue
            # Skip category rows like "ข้อมูลผิด", "ขาดข้อมูล", "เพิ่มรายละเอียด"
            if name in ('ข้อมูลผิด / ขาดข้อจำกัดสำคัญ', 'ขาดข้อมูลพืช/อัตรา', 'เพิ่มรายละเอียด'):
                continue

            update = {'product_name': name, '_sheet': sheet_name}

            # Read columns that may have updated data
            crops = clean_text(ws.cell(r, 5).value)
            rate = clean_text(ws.cell(r, 6).value)
            pest = clean_text(ws.cell(r, pest_idx).value) if pest_idx else None
            caution = clean_text(ws.cell(r, 9).value)

            # how_to_use and usage_period column positions differ by sheet
            if 'PGR' in sheet_name:
                how = clean_text(ws.cell(r, 10).value)
                period = clean_text(ws.cell(r, 11).value)
            else:
                how = clean_text(ws.cell(r, 10).value)
                period = clean_text(ws.cell(r, 11).value)

            if crops:
                update['applicable_crops'] = crops
            if rate:
                update['usage_rate'] = rate
            if pest and pest_col:
                update[pest_col] = pest
            if caution:
                update['caution_notes'] = caution
            if how:
                update['how_to_use'] = how
            if period:
                update['usage_period'] = period

            updates.append(update)

    return updates


def apply_updates(updates):
    """Apply updates to Supabase products3."""
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    success = 0
    skipped = 0
    errors = 0

    for u in updates:
        name = u['product_name']
        sheet = u.pop('_sheet')

        # Build update dict (only non-None fields, exclude product_name)
        data = {k: v for k, v in u.items() if k != 'product_name' and v is not None}

        if not data:
            print(f"  [SKIP] {name} ({sheet}) — no data to update")
            skipped += 1
            continue

        try:
            # Find product by name (case-insensitive)
            existing = sb.table(PRODUCT_TABLE).select('id, product_name').ilike('product_name', f'%{name.strip()}%').execute()

            if not existing.data:
                print(f"  [NOT FOUND] {name} ({sheet})")
                errors += 1
                continue

            # Use exact match if possible
            exact = [p for p in existing.data if p['product_name'].strip() == name.strip()]
            if exact:
                pid = exact[0]['id']
                pname = exact[0]['product_name']
            else:
                pid = existing.data[0]['id']
                pname = existing.data[0]['product_name']

            sb.table(PRODUCT_TABLE).update(data).eq('id', pid).execute()

            cols = list(data.keys())
            print(f"  [OK] {pname} — updated: {', '.join(cols)}")
            success += 1

        except Exception as e:
            print(f"  [ERROR] {name} — {e}")
            errors += 1

    return success, skipped, errors


if __name__ == "__main__":
    print("=" * 60)
    print("Update products3 from feedback Excel")
    print("=" * 60)

    print("\n1. Reading Excel...")
    updates = read_excel()
    print(f"   Found {len(updates)} product updates")

    print("\n2. Applying updates to Supabase...")
    s, sk, e = apply_updates(updates)

    print("\n" + "=" * 60)
    print(f"Done: {s} updated, {sk} skipped, {e} errors")
    print("=" * 60)
